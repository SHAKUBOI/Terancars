import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Création de données simulées pour plusieurs réseaux sociaux
data = {
    "Réseau": ["Instagram", "Instagram", "Instagram", "Facebook", "Facebook", "Facebook", "LinkedIn", "LinkedIn", "LinkedIn", "Twitter", "Twitter", "Twitter"],
    "Mois": ["Janvier", "Février", "Mars", "Janvier", "Février", "Mars", "Janvier", "Février", "Mars", "Janvier", "Février", "Mars"],
    "Abonnés": [15000, 16200, 17500, 25000, 26500, 28000, 8000, 8500, 9200, 12000, 12800, 13500],
    "Nouveaux Abonnés": [1200, 1300, 1400, 1500, 1600, 1800, 500, 600, 700, 800, 900, 1000],
    "Publications": [30, 28, 32, 25, 22, 28, 15, 18, 20, 45, 42, 48],
    "Engagements": [4500, 4800, 5200, 3800, 4000, 4400, 1200, 1400, 1600, 2200, 2400, 2600],
    "Taux Engagement (%)": [3.2, 3.4, 3.5, 2.8, 2.9, 3.0, 2.5, 2.6, 2.8, 2.2, 2.3, 2.4],
    "Portée": [45000, 48000, 52000, 65000, 68000, 72000, 25000, 27000, 29000, 35000, 37000, 40000],
    "Clics Site Web": [800, 850, 900, 1200, 1300, 1400, 400, 450, 500, 300, 350, 400],
    "Messages Reçus": [150, 165, 180, 250, 275, 300, 80, 90, 100, 120, 135, 150]
}

df = pd.DataFrame(data)

# Création du fichier Excel avec plusieurs feuilles d'analyse
with pd.ExcelWriter('rapport_reseaux_sociaux.xlsx', engine='openpyxl') as writer:
    # 1. Données brutes
    df.to_excel(writer, sheet_name='Données Détaillées', index=False)
    
    # 2. Résumé par réseau social (derniers chiffres)
    resume_reseaux = df.groupby('Réseau').agg({
        'Abonnés': 'last',
        'Nouveaux Abonnés': 'sum',
        'Publications': 'sum',
        'Engagements': 'sum',
        'Taux Engagement (%)': 'mean',
        'Portée': 'sum',
        'Clics Site Web': 'sum',
        'Messages Reçus': 'sum'
    }).round(2)
    resume_reseaux.to_excel(writer, sheet_name='Résumé par Réseau')
    
    # 3. Analyse de croissance
    croissance = pd.DataFrame()
    for reseau in df['Réseau'].unique():
        data_reseau = df[df['Réseau'] == reseau]
        croissance_abonnes = ((data_reseau['Abonnés'].iloc[-1] / data_reseau['Abonnés'].iloc[0] - 1) * 100).round(2)
        croissance_engagement = ((data_reseau['Taux Engagement (%)'].iloc[-1] / data_reseau['Taux Engagement (%)'].iloc[0] - 1) * 100).round(2)
        croissance = pd.concat([croissance, pd.DataFrame({
            'Réseau': [reseau],
            'Croissance Abonnés (%)': [croissance_abonnes],
            'Croissance Engagement (%)': [croissance_engagement],
            'Total Nouveaux Abonnés': [data_reseau['Nouveaux Abonnés'].sum()],
            'Moyenne Publications/Mois': [data_reseau['Publications'].mean()],
            'Moyenne Clics/Mois': [data_reseau['Clics Site Web'].mean()]
        })])
    croissance.to_excel(writer, sheet_name='Analyse Croissance', index=False)
    
    # 4. KPIs principaux
    kpis = pd.DataFrame({
        'Métrique': [
            'Total Abonnés',
            'Total Nouveaux Abonnés',
            'Total Publications',
            'Moyenne Taux Engagement',
            'Total Portée',
            'Total Clics Site Web',
            'Total Messages Reçus'
        ],
        'Valeur': [
            df.groupby('Réseau')['Abonnés'].last().sum(),
            df['Nouveaux Abonnés'].sum(),
            df['Publications'].sum(),
            df['Taux Engagement (%)'].mean(),
            df['Portée'].sum(),
            df['Clics Site Web'].sum(),
            df['Messages Reçus'].sum()
        ]
    })
    kpis.to_excel(writer, sheet_name='KPIs Globaux', index=False)
    
    # Formatage
    for sheet_name in writer.sheets:
        worksheet = writer.sheets[sheet_name]
        
        # Style d'en-tête
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Appliquer le style aux en-têtes
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Ajuster la largeur des colonnes
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

print("Rapport des réseaux sociaux créé avec succès !")
