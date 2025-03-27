import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Données initiales
data = {
    "Plateforme": ["Google Ads", "Google Ads", "Google Ads", "Facebook Ads", "Facebook Ads", "Facebook Ads"],
    "Campagne": ["Campagne 1", "Campagne 2", "Campagne 3", "Campagne 1", "Campagne 2", "Campagne 3"],
    "Impressions": [10000, 15000, 12000, 8000, 9500, 11000],
    "Clics": [500, 700, 600, 400, 450, 550],
    "CTR (%)": [5.0, 4.7, 5.0, 5.0, 4.7, 5.0],
    "CPC ($)": [0.50, 0.45, 0.55, 0.40, 0.45, 0.50],
    "Conversions": [50, 70, 60, 40, 45, 55],
    "CPA ($)": [10.00, 9.00, 11.00, 10.00, 9.50, 9.00],
    "ROI (%)": [500, 550, 520, 480, 500, 530]
}

df = pd.DataFrame(data)

# Créer un writer Excel avec le moteur openpyxl
with pd.ExcelWriter('rapport_performances_publicitaires.xlsx', engine='openpyxl') as writer:
    # 1. Feuille principale avec toutes les données
    df.to_excel(writer, sheet_name='Données Brutes', index=False)
    
    # 2. Feuille avec résumé par plateforme
    resume_plateforme = df.groupby('Plateforme').agg({
        'Impressions': 'sum',
        'Clics': 'sum',
        'CTR (%)': 'mean',
        'CPC ($)': 'mean',
        'Conversions': 'sum',
        'CPA ($)': 'mean',
        'ROI (%)': 'mean'
    }).round(2)
    resume_plateforme.to_excel(writer, sheet_name='Résumé Plateformes')
    
    # 3. Feuille avec les meilleures performances
    meilleures_performances = pd.DataFrame({
        'Métrique': [
            'Meilleur CTR',
            'Meilleur CPC',
            'Meilleur CPA',
            'Meilleur ROI'
        ],
        'Valeur': [
            f"{df['CTR (%)'].max()}%",
            f"${df['CPC ($)'].min()}",
            f"${df['CPA ($)'].min()}",
            f"{df['ROI (%)'].max()}%"
        ],
        'Plateforme': [
            df.loc[df['CTR (%)'].idxmax(), 'Plateforme'],
            df.loc[df['CPC ($)'].idxmin(), 'Plateforme'],
            df.loc[df['CPA ($)'].idxmin(), 'Plateforme'],
            df.loc[df['ROI (%)'].idxmax(), 'Plateforme']
        ],
        'Campagne': [
            df.loc[df['CTR (%)'].idxmax(), 'Campagne'],
            df.loc[df['CPC ($)'].idxmin(), 'Campagne'],
            df.loc[df['CPA ($)'].idxmin(), 'Campagne'],
            df.loc[df['ROI (%)'].idxmax(), 'Campagne']
        ]
    })
    meilleures_performances.to_excel(writer, sheet_name='Meilleures Performances', index=False)
    
    # Récupérer le classeur pour le formatage
    workbook = writer.book
    
    # Formatage de la feuille principale
    worksheet = writer.sheets['Données Brutes']
    
    # Définir les styles
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Appliquer le formatage aux en-têtes
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Formater les colonnes numériques
    number_columns = ['C', 'D', 'E', 'F', 'G', 'H', 'I']  # Colonnes avec des nombres
    for col in number_columns:
        for cell in worksheet[col][1:]:
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right')
            cell.border = border
    
    # Ajuster la largeur des colonnes
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

print("Rapport Excel créé avec succès !")
